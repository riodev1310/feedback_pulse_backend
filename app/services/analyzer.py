from __future__ import annotations

import os
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from statistics import mean
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from app.core.config import settings
from app.services.openai_sentiment import label_vietnamese_sentiment_dataframe
from app.services.trained_model_sentiment import (
    label_vietnamese_sentiment_dataframe_with_trained_model,
)


POSITIVE_KEYWORDS = [
    "nhiet tinh",
    "tan tam",
    "tam huyet",
    "de hieu",
    "than thien",
    "ho tro",
    "truyen dat tot",
    "giai thich ro",
    "tiet hoc vui",
    "co tam",
    "chuyen nghiep",
    "dung gio",
    "hai long",
    "tot",
    "hieu bai",
    "co ich",
    "chi tiet",
]

NEGATIVE_KEYWORDS = [
    "kho hieu",
    "qua loa",
    "khong ro",
    "khong hieu",
    "khong dung trong tam",
    "khong nhac",
    "tre gio",
    "tri hoan",
    "thieu tai lieu",
    "it thuc hanh",
    "khong ho tro",
    "khong nhiet tinh",
    "khong tan tam",
    "yeu",
    "bat cap",
    "lang man",
    "qua nhanh",
    "qua cham",
    "khong cong bang",
    "kho nghe",
    "khong day",
    "khong day dung",
    "khong day du",
    "chua du",
    "bo qua",
    "khong thong bao",
    "khong tra loi",
    "khong than thien",
]

TRIVIAL_COMMENTS = {
    "",
    ".",
    "..",
    "...",
    "-",
    "--",
    "khong",
    "khong co",
    "khong co y kien",
    "khong co gi",
    "ko",
    "ko co",
    "k",
    "none",
    "n a",
    "na",
    "null",
    "no",
}

ASPECTS = {
    "delivery": {
        "label": "Truyền đạt",
        "keywords": [
            "truyen dat",
            "giai thich",
            "de hieu",
            "kho hieu",
            "giong noi",
            "noi",
            "giang",
            "day",
            "thuyet trinh",
        ],
        "positive": "Sinh viên đánh giá cao khả năng truyền đạt và cách giải thích của giảng viên.",
        "negative": "Cần cải thiện cách truyền đạt, diễn giải trọng tâm và tốc độ lên lớp để sinh viên dễ theo dõi hơn.",
    },
    "content": {
        "label": "Nội dung kiến thức",
        "keywords": [
            "kien thuc",
            "noi dung",
            "trong tam",
            "cot loi",
            "ly thuyet",
            "chuong",
            "chuyen de",
            "thieu",
            "day du",
        ],
        "positive": "Phản hồi tích cực tập trung vào việc nội dung học hữu ích và bám sát nhu cầu môn học.",
        "negative": "Nên rà soát lại độ bao phủ nội dung, nhấn mạnh các phần cốt lõi và bảo đảm kiến thức phục vụ đầu ra môn học.",
    },
    "practice": {
        "label": "Thực hành và bài tập",
        "keywords": [
            "thuc hanh",
            "bai tap",
            "lab",
            "project",
            "do an",
            "van dung",
            "thao tac",
            "vi du",
        ],
        "positive": "Sinh viên ghi nhận phần bài tập và thực hành có giá trị áp dụng.",
        "negative": "Cần tăng số lượng ví dụ, bài tập và hướng dẫn thực hành để gắn kiến thức với ứng dụng thực tế.",
    },
    "support": {
        "label": "Hỗ trợ và thái độ",
        "keywords": [
            "nhiet tinh",
            "tan tam",
            "tam huyet",
            "ho tro",
            "than thien",
            "tra loi",
            "hoi dap",
            "quan tam",
            "co tam",
        ],
        "positive": "Giảng viên được ghi nhận là nhiệt tình, thân thiện và hỗ trợ sinh viên tốt.",
        "negative": "Cần tăng mức độ phản hồi, giải đáp và chủ động hỗ trợ sinh viên trong quá trình học.",
    },
    "assessment": {
        "label": "Kiểm tra và chấm điểm",
        "keywords": [
            "diem",
            "cham",
            "thi",
            "kiem tra",
            "danh gia",
            "giua ky",
            "cuoi ky",
            "thu hoach",
            "de thi",
        ],
        "positive": "Khâu kiểm tra và phản hồi kết quả được một phần sinh viên nhìn nhận tích cực.",
        "negative": "Nên chuẩn hóa lịch kiểm tra, tiêu chí chấm điểm và truyền thông sớm hơn về yêu cầu đánh giá.",
    },
    "schedule": {
        "label": "Tổ chức và thời gian",
        "keywords": [
            "tre gio",
            "dung gio",
            "lich",
            "thoi gian",
            "tri hoan",
            "thong bao",
            "buoi",
            "sap xep",
            "ke hoach",
        ],
        "positive": "Khâu tổ chức lớp học nhìn chung ổn định ở những phản hồi tích cực.",
        "negative": "Cần tổ chức lịch học, thông báo và tiến độ môn học rõ ràng, đúng hẹn và nhất quán hơn.",
    },
    "language": {
        "label": "Ngôn ngữ và giao tiếp",
        "keywords": [
            "tieng anh",
            "phat am",
            "giao tiep",
            "ngon ngu",
            "kho nghe",
            "nghe",
            "noi nhanh",
            "noi cham",
        ],
        "positive": "Khả năng giao tiếp của giảng viên tạo được thiện cảm trong một phần phản hồi.",
        "negative": "Nên cải thiện độ rõ ràng trong giao tiếp và nhất quán hơn khi sử dụng ngôn ngữ giảng dạy.",
    },
    "materials": {
        "label": "Tài liệu học tập",
        "keywords": [
            "tai lieu",
            "slide",
            "file",
            "tai nguyen",
            "tham khao",
            "mo ta",
            "huong dan",
        ],
        "positive": "Tài liệu và hướng dẫn học tập được đánh giá là hữu ích trong các phản hồi tích cực.",
        "negative": "Cần bổ sung tài liệu, ví dụ minh họa và hướng dẫn để sinh viên tự học thuận lợi hơn.",
    },
}


def compact_text(text: Any) -> str:
    if text is None:
        return ""
    value = str(text).replace("\r", " ").replace("\n", " ").strip()
    value = re.sub(r"\s+", " ", value)
    return value


def normalize_text(text: Any) -> str:
    value = compact_text(text).lower()
    value = value.replace("đ", "d")
    value = unicodedata.normalize("NFD", value)
    value = "".join(ch for ch in value if unicodedata.category(ch) != "Mn")
    value = re.sub(r"[^a-z0-9\s]", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def to_int(value: Any) -> int | None:
    numeric = to_float(value)
    if numeric is None:
        return None
    return int(numeric)


def is_informative_comment(text: str) -> bool:
    normalized = normalize_text(text)
    if normalized in TRIVIAL_COMMENTS:
        return False
    if len(normalized.split()) < 2:
        return False
    return len(normalized) >= 5


def sentiment_from_text(text: str, source_polarity: str) -> str:
    normalized = normalize_text(text)
    if not is_informative_comment(text):
        return "neutral"

    positive_hits = sum(1 for keyword in POSITIVE_KEYWORDS if keyword in normalized)
    negative_hits = sum(1 for keyword in NEGATIVE_KEYWORDS if keyword in normalized)

    score = positive_hits - negative_hits
    if source_polarity == "positive":
        score += 2
    elif source_polarity == "negative":
        score -= 2

    if positive_hits and negative_hits and abs(score) <= 1:
        return "neutral"
    if score >= 1:
        return "positive"
    if score <= -1:
        return "negative"
    return "neutral"


def label_records_with_sentiment(records: list[dict[str, Any]]) -> str:
    if not records:
        return "rule_based"

    if settings.sentiment_provider == "rule_based":
        for record in records:
            record["sentiment"] = sentiment_from_text(record["comment"], record["source_polarity"])
        return "rule_based"

    if settings.sentiment_provider not in {"openai", "trained_model", "pickle"}:
        raise ValueError(
            "Unsupported SENTIMENT_PROVIDER. Expected one of: rule_based, openai, trained_model."
        )

    unique_rows = []
    seen_comment_keys: set[str] = set()
    for record in records:
        comment_key = record["comment_key"]
        if comment_key in seen_comment_keys:
            continue
        seen_comment_keys.add(comment_key)
        unique_rows.append(
            {
                "comment_key": comment_key,
                "text": record["comment"],
            }
        )

    sentiments_df = pd.DataFrame(unique_rows)
    if settings.sentiment_provider in {"trained_model", "pickle"}:
        labeled_df = label_vietnamese_sentiment_dataframe_with_trained_model(sentiments_df)
        sentiment_engine = f"trained_model:{settings.trained_model_sentiment_model_path}"
    else:
        labeled_df = label_vietnamese_sentiment_dataframe(sentiments_df)
        sentiment_engine = f"openai:{settings.openai_sentiment_model}"
    label_map = dict(zip(labeled_df["comment_key"], labeled_df["label"], strict=False))

    for record in records:
        record["sentiment"] = label_map.get(
            record["comment_key"],
            sentiment_from_text(record["comment"], record["source_polarity"]),
        )

    return sentiment_engine


def detect_aspects(text: str) -> list[str]:
    normalized = normalize_text(text)
    hits: list[tuple[int, str]] = []
    for aspect_id, config in ASPECTS.items():
        count = sum(1 for keyword in config["keywords"] if keyword in normalized)
        if count:
            hits.append((count, aspect_id))
    hits.sort(key=lambda item: (-item[0], ASPECTS[item[1]]["label"]))
    if not hits:
        return ["general"]
    return [aspect_id for _, aspect_id in hits[:3]]


def preview_text(text: str, limit: int = 220) -> str:
    value = compact_text(text)
    if len(value) <= limit:
        return value
    return value[: limit - 1].rstrip() + "…"


def percentage(part: int, total: int) -> float:
    if not total:
        return 0.0
    return round(part * 100.0 / total, 1)


def canonical_lecturer_key(lecturer: Any, faculty: Any) -> str:
    return f"{normalize_text(lecturer)}|{normalize_text(faculty)}"


def pick_display_name(names: list[str]) -> str:
    cleaned = [compact_text(name) for name in names if compact_text(name)]
    if not cleaned:
        return ""
    counts = Counter(cleaned)
    return max(counts.items(), key=lambda item: (item[1], len(item[0]), item[0]))[0]


def build_feedback_records(workbook_path: str) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    records: list[dict[str, Any]] = []
    classes: dict[str, dict[str, Any]] = {}
    record_id = 1

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        current_meta: dict[str, Any] = {}
        response_bucket = ">=50%" if ">= 50%" in sheet_name else "<50%"

        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value is not None and str(value).strip() for value in row):
                continue

            if row[1] is not None:
                current_meta = {
                    "lecturer": compact_text(row[1]),
                    "faculty": compact_text(row[2]),
                    "course": compact_text(row[3]),
                    "program": compact_text(row[4]),
                    "class_name": compact_text(row[5]),
                    "class_size": to_int(row[6]),
                    "participants": to_int(row[7]),
                    "feedback_count": to_int(row[8]),
                    "average_score": to_float(row[9]),
                }
            elif current_meta:
                updates = {
                    "faculty": compact_text(row[2]) if row[2] is not None else current_meta.get("faculty", ""),
                    "course": compact_text(row[3]) if row[3] is not None else current_meta.get("course", ""),
                    "program": compact_text(row[4]) if row[4] is not None else current_meta.get("program", ""),
                    "class_name": compact_text(row[5]) if row[5] is not None else current_meta.get("class_name", ""),
                }
                current_meta.update(updates)

            if not current_meta.get("lecturer"):
                continue

            class_key = "|".join(
                [
                    current_meta.get("lecturer", ""),
                    current_meta.get("faculty", ""),
                    current_meta.get("course", ""),
                    current_meta.get("program", ""),
                    current_meta.get("class_name", ""),
                ]
            )
            if class_key and class_key not in classes:
                classes[class_key] = {
                    **current_meta,
                    "class_key": class_key,
                    "response_bucket": response_bucket,
                }

            for column_index, source_polarity in ((10, "positive"), (11, "negative")):
                comment = compact_text(row[column_index])
                if not is_informative_comment(comment):
                    continue

                records.append(
                    {
                        "record_id": record_id,
                        "sheet_name": sheet_name,
                        "row_number": row_number,
                        "source_polarity": source_polarity,
                        "sentiment": "neutral",
                        "comment": comment,
                        "comment_key": normalize_text(comment),
                        "aspects": detect_aspects(comment),
                        "response_bucket": response_bucket,
                        "class_key": class_key,
                        **current_meta,
                    }
                )
                record_id += 1

    return records, classes


def aspect_counter(records: list[dict[str, Any]], sentiment: str | None = None) -> Counter:
    counter: Counter = Counter()
    for record in records:
        if sentiment and record["sentiment"] != sentiment:
            continue
        for aspect_id in record["aspects"]:
            counter[aspect_id] += 1
    return counter


def build_comment_preview(records: list[dict[str, Any]], limit: int = 10) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for record in records:
        key = record["comment_key"]
        if key not in grouped:
            grouped[key] = {
                "comment": record["comment"],
                "sentiment": record["sentiment"],
                "source_polarity": record["source_polarity"],
                "count": 0,
                "aspects": Counter(),
            }
        grouped[key]["count"] += 1
        grouped[key]["aspects"].update(record["aspects"])

    previews: list[dict[str, Any]] = []
    for item in grouped.values():
        top_aspects = [ASPECTS[aspect]["label"] for aspect, _ in item["aspects"].most_common(2) if aspect in ASPECTS]
        previews.append(
            {
                "comment": preview_text(item["comment"]),
                "sentiment": item["sentiment"],
                "source_polarity": item["source_polarity"],
                "count": item["count"],
                "aspects": top_aspects or ["Khác"],
            }
        )

    previews.sort(key=lambda item: (-item["count"], item["sentiment"], item["comment"]))
    return previews[:limit]


def build_highlights(records: list[dict[str, Any]], sentiment: str, limit: int = 3) -> list[dict[str, Any]]:
    counter = aspect_counter(records, sentiment)
    highlights: list[dict[str, Any]] = []
    for aspect_id, count in counter.most_common(limit):
        if aspect_id not in ASPECTS:
            continue
        config = ASPECTS[aspect_id]
        highlights.append(
            {
                "aspect_id": aspect_id,
                "label": config["label"],
                "count": count,
                "message": config["positive"] if sentiment == "positive" else config["negative"],
            }
        )
    return highlights


def build_suggestions(records: list[dict[str, Any]], limit: int = 3) -> list[dict[str, Any]]:
    negative_records = [record for record in records if record["sentiment"] == "negative"]
    counter = aspect_counter(negative_records)
    suggestions: list[dict[str, Any]] = []
    for aspect_id, count in counter.most_common(limit):
        if aspect_id not in ASPECTS:
            continue
        config = ASPECTS[aspect_id]
        related = [record["comment"] for record in negative_records if aspect_id in record["aspects"]][:2]
        suggestions.append(
            {
                "label": config["label"],
                "count": count,
                "recommendation": config["negative"],
                "evidence": [preview_text(comment, 140) for comment in related],
            }
        )
    return suggestions


def narrative_from_score(score: float | None, positive_ratio: float, negative_ratio: float) -> str:
    if score is None:
        if positive_ratio >= 55:
            return "Tổng thể phản hồi nghiêng tích cực, nhưng vẫn còn một số điểm cần theo dõi."
        if negative_ratio >= 45:
            return "Tổng thể phản hồi cho thấy nhiều tín hiệu cần can thiệp ưu tiên."
        return "Phản hồi đang ở mức trung tính và cần thêm dữ liệu để kết luận chắc chắn."

    if score >= 3.6 and positive_ratio >= 55:
        return "Đây là hồ sơ giảng dạy mạnh, phản hồi tích cực chiếm ưu thế và điểm đánh giá lớp học ở mức cao."
    if score >= 3.2 and negative_ratio < 40:
        return "Chất lượng phản hồi nhìn chung ổn định; cần tinh chỉnh một vài điểm nổi bật để nâng trải nghiệm học tập."
    if negative_ratio >= 45 or score < 3.0:
        return "Giảng viên đang có tín hiệu cảnh báo về trải nghiệm lớp học; nên ưu tiên xử lý các vấn đề lặp lại trong học kỳ kế tiếp."
    return "Phản hồi cho thấy chất lượng ở mức trung bình khá, với cả điểm mạnh và điểm nghẽn cùng xuất hiện."


def summarize_lecturer(
    lecturer: str,
    lecturer_records: list[dict[str, Any]],
    lecturer_classes: list[dict[str, Any]],
    lecturer_key: str,
) -> dict[str, Any]:
    sentiment_counts = Counter(record["sentiment"] for record in lecturer_records)
    total_records = len(lecturer_records)
    positive_ratio = percentage(sentiment_counts["positive"], total_records)
    negative_ratio = percentage(sentiment_counts["negative"], total_records)
    neutral_ratio = percentage(sentiment_counts["neutral"], total_records)

    scores = [item["average_score"] for item in lecturer_classes if item.get("average_score") is not None]
    avg_score = round(mean(scores), 2) if scores else None
    total_class_size = sum(item.get("class_size") or 0 for item in lecturer_classes)
    total_participants = sum(item.get("participants") or 0 for item in lecturer_classes)

    faculties = sorted({item.get("faculty", "") for item in lecturer_classes if item.get("faculty")})
    programs = sorted({item.get("program", "") for item in lecturer_classes if item.get("program")})
    courses = sorted({item.get("course", "") for item in lecturer_classes if item.get("course")})

    top_strengths = build_highlights(lecturer_records, "positive")
    top_issues = build_highlights(lecturer_records, "negative")
    suggestions = build_suggestions(lecturer_records)

    summary = {
        "key_strengths": [item["message"] for item in top_strengths] or ["Chưa có đủ phản hồi tích cực lặp lại để rút ra điểm mạnh nổi trội."],
        "areas_for_improvement": [item["message"] for item in top_issues] or ["Chưa phát hiện nhóm vấn đề tiêu cực nổi trội trong dữ liệu hiện tại."],
        "overall_assessment": narrative_from_score(avg_score, positive_ratio, negative_ratio),
    }

    class_cards = []
    for item in sorted(
        lecturer_classes,
        key=lambda row: ((row.get("average_score") is None), row.get("average_score") or 0.0),
    ):
        participation_rate = percentage(item.get("participants") or 0, item.get("class_size") or 0)
        class_cards.append(
            {
                "class_name": item.get("class_name", ""),
                "course": item.get("course", ""),
                "program": item.get("program", ""),
                "average_score": item.get("average_score"),
                "feedback_count": item.get("feedback_count"),
                "participation_rate": participation_rate,
                "response_bucket": item.get("response_bucket"),
            }
        )

    return {
        "id": lecturer_key,
        "lecturer": lecturer,
        "faculty": faculties[0] if faculties else "",
        "faculties": faculties,
        "programs": programs,
        "courses": courses,
        "classes_taught": len(lecturer_classes),
        "feedback_snippets": total_records,
        "average_score": avg_score,
        "participation_rate": percentage(total_participants, total_class_size),
        "sentiment": {
            "positive": sentiment_counts["positive"],
            "negative": sentiment_counts["negative"],
            "neutral": sentiment_counts["neutral"],
            "positive_ratio": positive_ratio,
            "negative_ratio": negative_ratio,
            "neutral_ratio": neutral_ratio,
        },
        "top_strengths": top_strengths,
        "top_issues": top_issues,
        "suggestions": suggestions,
        "summary": summary,
        "comments_preview": build_comment_preview(lecturer_records),
        "classes": class_cards[:8],
    }


def build_faculty_overview(lecturer_summaries: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for summary in lecturer_summaries:
        grouped[summary["faculty"] or "Chưa phân khoa"].append(summary)

    overview = []
    for faculty, items in grouped.items():
        score_values = [item["average_score"] for item in items if item["average_score"] is not None]
        positives = sum(item["sentiment"]["positive"] for item in items)
        negatives = sum(item["sentiment"]["negative"] for item in items)
        neutrals = sum(item["sentiment"]["neutral"] for item in items)
        overview.append(
            {
                "faculty": faculty,
                "lecturer_count": len(items),
                "class_count": sum(item["classes_taught"] for item in items),
                "average_score": round(mean(score_values), 2) if score_values else None,
                "positive": positives,
                "negative": negatives,
                "neutral": neutrals,
                "negative_ratio": percentage(negatives, positives + negatives + neutrals),
            }
        )
    overview.sort(key=lambda item: (-item["negative_ratio"], item["faculty"]))
    return overview


def build_institution_report(
    workbook_name: str,
    records: list[dict[str, Any]],
    classes: dict[str, dict[str, Any]],
    lecturer_summaries: list[dict[str, Any]],
    sentiment_engine: str,
) -> dict[str, Any]:
    sentiment_counts = Counter(record["sentiment"] for record in records)
    total_class_size = sum(item.get("class_size") or 0 for item in classes.values())
    total_participants = sum(item.get("participants") or 0 for item in classes.values())
    score_values = [item["average_score"] for item in classes.values() if item.get("average_score") is not None]
    top_negative_aspects = build_highlights(records, "negative", limit=4)
    top_positive_aspects = build_highlights(records, "positive", limit=4)

    at_risk = sorted(
        lecturer_summaries,
        key=lambda item: (
            item["average_score"] is None,
            item["average_score"] if item["average_score"] is not None else 99,
            -item["sentiment"]["negative_ratio"],
        ),
    )[:5]

    return {
        "workbook_name": workbook_name,
        "analyzed_at": datetime.now().isoformat(timespec="seconds"),
        "sentiment_engine": sentiment_engine,
        "lecturer_count": len(lecturer_summaries),
        "class_count": len(classes),
        "feedback_snippets": len(records),
        "average_score": round(mean(score_values), 2) if score_values else None,
        "participation_rate": percentage(total_participants, total_class_size),
        "sentiment": {
            "positive": sentiment_counts["positive"],
            "negative": sentiment_counts["negative"],
            "neutral": sentiment_counts["neutral"],
            "positive_ratio": percentage(sentiment_counts["positive"], len(records)),
            "negative_ratio": percentage(sentiment_counts["negative"], len(records)),
            "neutral_ratio": percentage(sentiment_counts["neutral"], len(records)),
        },
        "faculties": build_faculty_overview(lecturer_summaries),
        "top_negative_aspects": top_negative_aspects,
        "top_positive_aspects": top_positive_aspects,
        "priority_lecturers": [
            {
                "lecturer": item["lecturer"],
                "faculty": item["faculty"],
                "average_score": item["average_score"],
                "negative_ratio": item["sentiment"]["negative_ratio"],
            }
            for item in at_risk
        ],
    }


def analyze_workbook(workbook_path: str, workbook_name: str | None = None) -> dict[str, Any]:
    records, classes = build_feedback_records(workbook_path)
    display_name = workbook_name or os.path.basename(workbook_path)
    sentiment_engine = label_records_with_sentiment(records)

    grouped_records: dict[str, list[dict[str, Any]]] = defaultdict(list)
    grouped_classes: dict[str, list[dict[str, Any]]] = defaultdict(list)
    grouped_names: dict[str, list[str]] = defaultdict(list)

    for record in records:
        lecturer_key = canonical_lecturer_key(record.get("lecturer"), record.get("faculty"))
        grouped_records[lecturer_key].append(record)
        grouped_names[lecturer_key].append(record.get("lecturer", ""))

    for class_item in classes.values():
        lecturer_key = canonical_lecturer_key(class_item.get("lecturer"), class_item.get("faculty"))
        grouped_classes[lecturer_key].append(class_item)
        grouped_names[lecturer_key].append(class_item.get("lecturer", ""))

    lecturer_summaries = [
        summarize_lecturer(
            pick_display_name(grouped_names[lecturer_key]),
            grouped_records[lecturer_key],
            grouped_classes.get(lecturer_key, []),
            lecturer_key,
        )
        for lecturer_key in grouped_records
    ]
    lecturer_summaries.sort(
        key=lambda item: (
            item["average_score"] is None,
            -(item["average_score"] or 0),
            item["lecturer"],
        )
    )

    report = build_institution_report(display_name, records, classes, lecturer_summaries, sentiment_engine)
    return {
        "report": report,
        "lecturers": lecturer_summaries,
    }
